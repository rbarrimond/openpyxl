# from openpyxl import Workbook
# from openpyxl.chart import BarChart, Reference

# workbook = Workbook()
# sheet = workbook.active

# # sample data
# rows = [
#     ["Product", "Online", "Store"],
#     [1, 30, 45],
#     [2, 40, 30],
#     [3, 40, 25],
#     [4, 50, 50],
#     [5, 30, 25],
#     [6, 25, 35],
#     [7, 20, 40],
# ]

# for row in rows:
#     sheet.append(row)

# chart = BarChart()
# data = Reference(worksheet=sheet,
#                  min_row=1,
#                  max_row=8,
#                  min_col=2,
#                  max_col=3)

# chart.add_data(data, titles_from_data=True)
# sheet.add_chart(chart, "E2")

# workbook.save("chart.xlsx")


######## Line Chart
import random
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

workbook = Workbook()
sheet = workbook.active

# sample data
rows = [
    ["", "January", "February", "March", "April",
    "May", "June", "July", "August", "September",
    "October", "November", "December"],
    [1, ],
    [2, ],
    [3, ],
]

for row in rows:
    sheet.append(row)

for row in sheet.iter_rows(min_row=2,
                           max_row=4,
                           min_col=2,
                           max_col=13):
    for cell in row:
        cell.value = random.randrange(5, 100)

## Make chart

chart = LineChart()
data = Reference(worksheet=sheet,
                 min_row=2,
                 max_row=4,
                 min_col=1,
                 max_col=13)

chart.add_data(data, from_rows=True, titles_from_data=True)

# add categories
cats = Reference(worksheet=sheet,
                 min_row=1,
                 max_row=1,
                 min_col=2,
                 max_col=13)
chart.set_categories(cats)

# add axis labels
chart.x_axis.title = "Months"
chart.y_axis.title = "Sales (per unit)"

sheet.add_chart(chart, "C6")

workbook.save("line_chart.xlsx")