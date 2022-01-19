import json
from pprint import pp
from openpyxl import load_workbook

workbook = load_workbook(filename='openpyxl_sample_code/sample.xlsx')
sheet = workbook.active

products = {}

for row in sheet.iter_rows(min_row=2,
                           min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3],
    }
    
    products[product_id] = product
    
pp(products)
